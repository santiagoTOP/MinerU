import argparse
import atexit
import configparser
import copy
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from abc import ABC, abstractmethod
from contextlib import contextmanager
from pathlib import Path
from typing import List, Optional, Tuple, Union
from urllib.parse import parse_qs, unquote, urlparse

import chardet
import openpyxl
import pandas as pd
import requests
from docx import Document
from docx.oxml import CT_P, CT_Tbl
from docx.oxml.ns import qn
from docx.table import Table

# Copyright (c) Opendatalab. All rights reserved.
from loguru import logger
from openpyxl.utils import get_column_letter
from PIL import Image

from mineru.backend.pipeline.model_json_to_middle_json import (
    result_to_middle_json as pipeline_result_to_middle_json,
)
from mineru.backend.pipeline.pipeline_analyze import doc_analyze as pipeline_doc_analyze
from mineru.backend.pipeline.pipeline_middle_json_mkcontent import (
    union_make as pipeline_union_make,
)
from mineru.backend.vlm.vlm_analyze import doc_analyze as vlm_doc_analyze
from mineru.backend.vlm.vlm_middle_json_mkcontent import union_make as vlm_union_make
from mineru.cli.common import (
    convert_pdf_bytes_to_bytes_by_pypdfium2,
    prepare_env,
    read_fn,
)
from mineru.data.data_reader_writer import FileBasedDataWriter
from mineru.utils.draw_bbox import draw_layout_bbox, draw_span_bbox
from mineru.utils.enum_class import MakeMode
from mineru.utils.guess_suffix_or_lang import guess_suffix_by_path


class Config:
    """配置管理类"""
    def __init__(self, config_file: str = 'config.ini'):
        self.config = configparser.ConfigParser()
        # 设置默认值
        self.config['DEFAULT'] = {
            'download_dir': '/home/hp-2/Agent_Platform/MinerU/demo/pdfs',
            'output_dir': '/home/hp-2/Agent_Platform/MinerU/output',
            'max_file_size_mb': '100',
            'timeout_seconds': '300'
        }
        
        # 尝试读取配置文件
        if os.path.exists(config_file):
            self.config.read(config_file)
    
    def get_download_dir(self) -> str:
        return os.path.expanduser(self.config.get('DEFAULT', 'download_dir'))
    
    def get_output_dir(self) -> str:
        return os.path.expanduser(self.config.get('DEFAULT', 'output_dir'))
    
    def get_max_file_size_mb(self) -> int:
        return self.config.getint('DEFAULT', 'max_file_size_mb')
    
    def get_timeout_seconds(self) -> int:
        return self.config.getint('DEFAULT', 'timeout_seconds')


class TempFileManager:
    """临时文件管理器"""
    def __init__(self):
        self.temp_files = []
        self.temp_dirs = []
    
    def add_temp_file(self, file_path: str):
        """添加临时文件到清理列表"""
        self.temp_files.append(file_path)
    
    def add_temp_dir(self, dir_path: str):
        """添加临时目录到清理列表"""
        self.temp_dirs.append(dir_path)
    
    def cleanup(self):
        """清理所有临时文件和目录"""
        for file_path in self.temp_files:
            try:
                if os.path.exists(file_path):
                    os.unlink(file_path)
                    logger.debug(f"已清理临时文件: {file_path}")
            except Exception as e:
                logger.warning(f"清理临时文件失败 {file_path}: {e}")
        
        for dir_path in self.temp_dirs:
            try:
                if os.path.exists(dir_path):
                    shutil.rmtree(dir_path, ignore_errors=True)
                    logger.debug(f"已清理临时目录: {dir_path}")
            except Exception as e:
                logger.warning(f"清理临时目录失败 {dir_path}: {e}")


# 全局临时文件管理器
temp_manager = TempFileManager()
atexit.register(temp_manager.cleanup)


@contextmanager
def timer(operation_name: str):
    """计时器上下文管理器"""
    start_time = time.time()
    try:
        yield
    finally:
        elapsed = time.time() - start_time
        logger.info(f"{operation_name}耗时: {elapsed:.2f}秒")


def validate_file_size(file_path: str, max_size_mb: int = 100):
    """验证文件大小"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    file_size = os.path.getsize(file_path)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    if file_size > max_size_bytes:
        raise ValueError(f"文件过大: {file_size/1024/1024:.1f}MB > {max_size_mb}MB")
    
    logger.info(f"文件大小验证通过: {file_size/1024/1024:.1f}MB")


def read_text_file_with_encoding_detection(file_path: str) -> str:
    """使用编码检测读取文本文件"""
    try:
        # 首先尝试检测编码
        with open(file_path, "rb") as f:
            raw_data = f.read()
        
        detected = chardet.detect(raw_data)
        encoding = detected.get('encoding', 'utf-8')
        confidence = detected.get('confidence', 0)
        
        logger.info(f"检测到文件编码: {encoding} (置信度: {confidence:.2f})")
        
        # 定义编码尝试顺序
        # 如果检测到gb2312，优先尝试gbk（因为gbk是gb2312的超集）
        encoding_candidates = []
        if confidence >= 0.7:
            # 置信度高时，优先使用检测到的编码
            encoding_candidates.append(encoding)
            # 如果检测到gb2312，添加gbk作为备选（gbk包含gb2312）
            if encoding.lower() in ['gb2312', 'gb18030']:
                encoding_candidates.append('gbk')
            # 如果检测到gbk，添加gb18030作为备选
            elif encoding.lower() == 'gbk':
                encoding_candidates.append('gb18030')
        else:
            logger.warning("编码检测置信度较低，尝试常见编码")
        
        # 添加常见编码作为备选
        common_encodings = ['utf-8', 'gbk', 'gb18030', 'gb2312', 'utf-16', 'latin1']
        for enc in common_encodings:
            if enc not in encoding_candidates:
                encoding_candidates.append(enc)
        
        # 尝试使用各种编码读取文件
        last_error = None
        for enc in encoding_candidates:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    content = f.read()
                logger.info(f"成功使用编码 {enc} 读取文件")
                return content
            except (UnicodeDecodeError, UnicodeError) as e:
                last_error = e
                continue
        
        # 如果所有编码都失败，尝试使用errors='ignore'或errors='replace'
        logger.warning("所有编码尝试失败，使用errors='replace'模式")
        for enc in ['utf-8', 'gbk', 'latin1']:
            try:
                with open(file_path, "r", encoding=enc, errors='replace') as f:
                    content = f.read()
                logger.warning(f"使用编码 {enc} (errors='replace') 读取文件，部分字符可能丢失")
                return content
            except Exception:
                continue
        
        # 最后的备选方案：使用latin1（不会失败，但可能乱码）
        logger.error("所有编码尝试均失败，使用latin1编码（可能产生乱码）")
        with open(file_path, "r", encoding='latin1', errors='replace') as f:
            return f.read()
    
    except Exception as e:
        logger.error(f"读取文件失败: {e}")
        raise


def extract_attachment_id(url: str) -> str:
    """安全地从URL中提取attachment ID"""
    try:
        parsed = urlparse(url)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError("无效的URL格式")
        
        params = parse_qs(parsed.query)
        attachment_id = params.get('attachmentId', [None])[0]
        
        if not attachment_id:
            raise ValueError("URL中缺少attachmentId参数")
        
        # 验证attachment_id格式（只允许字母数字、连字符和下划线）
        if not re.match(r'^[a-zA-Z0-9\-_]+$', attachment_id):
            raise ValueError("无效的attachmentId格式")
        
        logger.info(f"成功提取attachment ID: {attachment_id}")
        return attachment_id
    
    except Exception as e:
        logger.error(f"URL解析失败: {e}")
        raise ValueError(f"URL解析失败: {e}")


def do_parse(
    output_dir,  # Output directory for storing parsing results
    pdf_file_names: list[str],  # List of PDF file names to be parsed
    pdf_bytes_list: list[bytes],  # List of PDF bytes to be parsed
    p_lang_list: list[str],  # List of languages for each PDF, default is 'ch' (Chinese)
    backend="pipeline",  # The backend for parsing PDF, default is 'pipeline'
    parse_method="auto",  # The method for parsing PDF, default is 'auto'
    formula_enable=True,  # Enable formula parsing
    table_enable=True,  # Enable table parsing
    server_url=None,  # Server URL for vlm-http-client backend
    f_draw_layout_bbox=False,  # Whether to draw layout bounding boxes
    f_draw_span_bbox=False,  # Whether to draw span bounding boxes
    f_dump_md=True,  # Whether to dump markdown files
    f_dump_middle_json=False,  # Whether to dump middle JSON files
    f_dump_model_output=False,  # Whether to dump model output files
    f_dump_orig_pdf=False,  # Whether to dump original PDF files
    f_dump_content_list=False,  # Whether to dump content list files
    f_make_md_mode=MakeMode.MM_MD,  # The mode for making markdown content, default is MM_MD
    start_page_id=0,  # Start page ID for parsing, default is 0
    end_page_id=None,  # End page ID for parsing, default is None (parse all pages until the end of the document)
):

    if backend == "pipeline":
        for idx, pdf_bytes in enumerate(pdf_bytes_list):
            new_pdf_bytes = convert_pdf_bytes_to_bytes_by_pypdfium2(pdf_bytes, start_page_id, end_page_id)
            pdf_bytes_list[idx] = new_pdf_bytes

        infer_results, all_image_lists, all_pdf_docs, lang_list, ocr_enabled_list = pipeline_doc_analyze(pdf_bytes_list, p_lang_list, parse_method=parse_method, formula_enable=formula_enable,table_enable=table_enable)

        for idx, model_list in enumerate(infer_results):
            model_json = copy.deepcopy(model_list)
            pdf_file_name = pdf_file_names[idx]
            local_image_dir, local_md_dir = prepare_env(output_dir, pdf_file_name, parse_method)
            image_writer, md_writer = FileBasedDataWriter(local_image_dir), FileBasedDataWriter(local_md_dir)

            images_list = all_image_lists[idx]
            pdf_doc = all_pdf_docs[idx]
            _lang = lang_list[idx]
            _ocr_enable = ocr_enabled_list[idx]
            middle_json = pipeline_result_to_middle_json(model_list, images_list, pdf_doc, image_writer, _lang, _ocr_enable, formula_enable)

            pdf_info = middle_json["pdf_info"]

            pdf_bytes = pdf_bytes_list[idx]
            _process_output(
                pdf_info, pdf_bytes, pdf_file_name, local_md_dir, local_image_dir,
                md_writer, f_draw_layout_bbox, f_draw_span_bbox, f_dump_orig_pdf,
                f_dump_md, f_dump_content_list, f_dump_middle_json, f_dump_model_output,
                f_make_md_mode, middle_json, model_json, is_pipeline=True
            )
    else:
        if backend.startswith("vlm-"):
            backend = backend[4:]

        f_draw_span_bbox = False
        parse_method = "vlm"
        for idx, pdf_bytes in enumerate(pdf_bytes_list):
            pdf_file_name = pdf_file_names[idx]
            pdf_bytes = convert_pdf_bytes_to_bytes_by_pypdfium2(pdf_bytes, start_page_id, end_page_id)
            local_image_dir, local_md_dir = prepare_env(output_dir, pdf_file_name, parse_method)
            image_writer, md_writer = FileBasedDataWriter(local_image_dir), FileBasedDataWriter(local_md_dir)
            middle_json, infer_result = vlm_doc_analyze(pdf_bytes, image_writer=image_writer, backend=backend, server_url=server_url)

            pdf_info = middle_json["pdf_info"]

            _process_output(
                pdf_info, pdf_bytes, pdf_file_name, local_md_dir, local_image_dir,
                md_writer, f_draw_layout_bbox, f_draw_span_bbox, f_dump_orig_pdf,
                f_dump_md, f_dump_content_list, f_dump_middle_json, f_dump_model_output,
                f_make_md_mode, middle_json, infer_result, is_pipeline=False
            )


def _process_output(
        pdf_info,
        pdf_bytes,
        pdf_file_name,
        local_md_dir,
        local_image_dir,
        md_writer,
        f_draw_layout_bbox,
        f_draw_span_bbox,
        f_dump_orig_pdf,
        f_dump_md,
        f_dump_content_list,
        f_dump_middle_json,
        f_dump_model_output,
        f_make_md_mode,
        middle_json,
        model_output=None,
        is_pipeline=True
):
    """处理输出文件"""
    if f_draw_layout_bbox:
        draw_layout_bbox(pdf_info, pdf_bytes, local_md_dir, f"{pdf_file_name}_layout.pdf")

    if f_draw_span_bbox:
        draw_span_bbox(pdf_info, pdf_bytes, local_md_dir, f"{pdf_file_name}_span.pdf")

    if f_dump_orig_pdf:
        md_writer.write(
            f"{pdf_file_name}_origin.pdf",
            pdf_bytes,
        )

    image_dir = str(os.path.basename(local_image_dir))

    if f_dump_md:
        make_func = pipeline_union_make if is_pipeline else vlm_union_make
        md_content_str = make_func(pdf_info, f_make_md_mode, image_dir)
        md_writer.write_string(
            f"{pdf_file_name}.md",
            md_content_str,
        )

    if f_dump_content_list:
        make_func = pipeline_union_make if is_pipeline else vlm_union_make
        content_list = make_func(pdf_info, MakeMode.CONTENT_LIST, image_dir)
        md_writer.write_string(
            f"{pdf_file_name}_content_list.json",
            json.dumps(content_list, ensure_ascii=False, indent=4),
        )

    if f_dump_middle_json:
        md_writer.write_string(
            f"{pdf_file_name}_middle.json",
            json.dumps(middle_json, ensure_ascii=False, indent=4),
        )

    if f_dump_model_output:
        md_writer.write_string(
            f"{pdf_file_name}_model.json",
            json.dumps(model_output, ensure_ascii=False, indent=4),
        )

    logger.info(f"local output dir is {local_md_dir}")


def parse_doc(
        path_list: list[Path],
        output_dir,
        lang="ch",
        backend="pipeline",
        method="auto",
        server_url=None,
        start_page_id=0,
        end_page_id=None
):
    """
        Parameter description:
        path_list: List of document paths to be parsed, can be PDF or image files.
        output_dir: Output directory for storing parsing results.
        lang: Language option, default is 'ch', optional values include['ch', 'ch_server', 'ch_lite', 'en', 'korean', 'japan', 'chinese_cht', 'ta', 'te', 'ka']。
            Input the languages in the pdf (if known) to improve OCR accuracy.  Optional.
            Adapted only for the case where the backend is set to "pipeline"
        backend: the backend for parsing pdf:
            pipeline: More general.
            vlm-transformers: More general.
            vlm-vllm-engine: Faster(engine).
            vlm-http-client: Faster(client).
            without method specified, pipeline will be used by default.
        method: the method for parsing pdf:
            auto: Automatically determine the method based on the file type.
            txt: Use text extraction method.
            ocr: Use OCR method for image-based PDFs.
            Without method specified, 'auto' will be used by default.
            Adapted only for the case where the backend is set to "pipeline".
        server_url: When the backend is `http-client`, you need to specify the server_url, for example:`http://127.0.0.1:30000`
        start_page_id: Start page ID for parsing, default is 0
        end_page_id: End page ID for parsing, default is None (parse all pages until the end of the document)
    """
    try:
        file_name_list = []
        pdf_bytes_list = []
        lang_list = []
        for path in path_list:
            file_name = str(Path(path).stem)
            pdf_bytes = read_fn(path)
            file_name_list.append(file_name)
            pdf_bytes_list.append(pdf_bytes)
            lang_list.append(lang)
        do_parse(
            output_dir=output_dir,
            pdf_file_names=file_name_list,
            pdf_bytes_list=pdf_bytes_list,
            p_lang_list=lang_list,
            backend=backend,
            parse_method=method,
            server_url=server_url,
            start_page_id=start_page_id,
            end_page_id=end_page_id
        )
    except Exception as e:
        logger.exception(e)


class FileDownloader:
    """文件下载器"""
    
    def __init__(self, download_dir: str = "./uploadFile", max_size_mb: int = 100, timeout: int = 300):
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.max_size_mb = max_size_mb
        self.timeout = timeout

    def get_file_extension_from_response(self, response) -> tuple[Optional[str], Optional[str]]:
        """从响应头中提取文件名和后缀"""
        try:
            content_disposition = response.headers.get('Content-Disposition', '')
            if 'filename=' in content_disposition:
                filename = content_disposition.split('filename=')[1]
                filename = unquote(filename.strip('"'))
                file_extension = filename.split('.')[-1].lower()
                logger.info(f"从响应头获取文件信息: {filename}, 扩展名: {file_extension}")
                return filename, file_extension
            return None, None
        except Exception as e:
            logger.warning(f"解析响应头失败: {e}")
            return None, None

    def download_file(self, url: str) -> str:
        """下载文件"""
        try:
            logger.info(f"开始下载文件: {url}")
            
            # 验证URL
            parsed_url = urlparse(url)
            if not parsed_url.scheme or not parsed_url.netloc:
                raise ValueError("无效的URL格式")
            
            # 发送请求获取文件
            response = requests.get(url, stream=True, timeout=self.timeout)
            response.raise_for_status()
            
            # 检查文件大小
            content_length = response.headers.get('content-length')
            if content_length:
                file_size_mb = int(content_length) / (1024 * 1024)
                if file_size_mb > self.max_size_mb:
                    raise ValueError(f"文件过大: {file_size_mb:.1f}MB > {self.max_size_mb}MB")
                logger.info(f"文件大小: {file_size_mb:.1f}MB")

            # 获取文件名和后缀
            file_name, file_extension = self.get_file_extension_from_response(response)
            
            # 提取attachment ID作为备用文件名
            try:
                attachment_id = extract_attachment_id(url)
            except ValueError as e:
                logger.warning(f"提取attachment ID失败: {e}")
                attachment_id = f"download_{uuid.uuid4().hex[:8]}"

            if not file_extension:
                # 尝试从URL路径中推断扩展名
                url_path = parsed_url.path
                if '.' in url_path:
                    file_extension = url_path.split('.')[-1].lower()
                else:
                    raise ValueError("无法确定文件扩展名")

            # 构建文件路径
            file_name = f"{attachment_id}.{file_extension}"
            file_path = self.download_dir / file_name
            
            # 确保文件名唯一
            counter = 1
            original_path = file_path
            while file_path.exists():
                stem = original_path.stem
                suffix = original_path.suffix
                file_path = original_path.parent / f"{stem}_{counter}{suffix}"
                counter += 1

            # 下载文件
            downloaded_size = 0
            with open(file_path, "wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        file.write(chunk)
                        downloaded_size += len(chunk)
                        
                        # 检查下载大小限制
                        if downloaded_size > self.max_size_mb * 1024 * 1024:
                            file.close()
                            file_path.unlink()  # 删除部分下载的文件
                            raise ValueError(f"下载文件过大，超过 {self.max_size_mb}MB 限制")

            logger.info(f"文件下载成功: {file_path}")
            temp_manager.add_temp_file(str(file_path))
            return str(file_path)

        except requests.exceptions.Timeout:
            logger.error("下载超时")
            raise RuntimeError("文件下载超时")
        except requests.exceptions.ConnectionError:
            logger.error("网络连接失败")
            raise RuntimeError("网络连接失败")
        except requests.exceptions.RequestException as e:
            logger.error(f"下载请求失败: {e}")
            raise RuntimeError(f"文件下载失败: {e}")
        except Exception as e:
            logger.error(f"下载过程出错: {e}")
            raise


class DocumentParser(ABC):
    """Abstract base class for document parsers"""
    def __init__(self):
        self.model_manager = None
    
    @abstractmethod
    def parse(self, file_bytes: bytes, filename: str, output_dir: Path, opts: dict):
        pass
    
    def _sanitize_filename(self, filename: str) -> str:
        return re.sub(r'[^\w\-_\. ]', '_', filename)
    
    def set_model_manager(self, model_manager):
        self.model_manager = model_manager


class WordParser(DocumentParser):
    def parse(self, file_path: str, output_dir: str) -> Optional[str]:
        """解析 Word 文档（支持 .doc 和 .docx），转换为 Markdown"""
        try:
            # 输入验证和目录创建
            self._validate_input(file_path)
            doc_dir, image_dir = self._create_output_dirs(file_path, output_dir)
            
            # 统一处理为.docx格式
            docx_path = self._convert_to_docx_if_needed(file_path, doc_dir)
            
            # 解析文档内容
            return self._parse_docx_content(docx_path, doc_dir, image_dir)
            
        except Exception as e:
            print(f"❌ 解析失败: {str(e)}")
            return None

    def _validate_input(self, file_path: str) -> None:
        """验证输入文件"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"输入文件不存在: {file_path}")
        if os.path.splitext(file_path)[1].lower() not in ('.doc', '.docx'):
            raise ValueError("仅支持 .doc 和 .docx 格式")

    def _create_output_dirs(self, file_path: str, output_dir: str) -> Tuple[str, str]:
        """创建输出目录结构"""
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        doc_dir = os.path.join(output_dir, base_name)
        image_dir = os.path.join(doc_dir, "images")
        os.makedirs(image_dir, exist_ok=True)
        return doc_dir, image_dir

    def _convert_to_docx_if_needed(self, file_path: str, output_dir: str) -> str:
        """将.doc转换为.docx（如需要）"""
        output_path = os.path.join(output_dir, os.path.basename(file_path).split(".")[0]) + '.docx'
        
        if file_path.lower().endswith('.doc'):
            try:
                user_install = f"file:///tmp/LibreOfficeProfile_{uuid.uuid4()}"
                subprocess.run([
                    "soffice", "--headless", f"-env:UserInstallation={user_install}", "--convert-to", "docx",
                    "--outdir", output_dir, file_path
                ], check=True, timeout=30)
                print(f"✅ 已转换.doc到.docx: {output_path}")
            except subprocess.TimeoutExpired:
                raise RuntimeError("文档转换超时，请确保已安装LibreOffice")
            except Exception as e:
                raise RuntimeError(f"文档转换失败: {str(e)}")
        else:
            shutil.copy2(file_path, output_path)
            
        return output_path

    def _parse_docx_content(self, file_path: str, doc_dir: str, image_dir: str) -> str:
        """解析.docx文档内容"""
        doc = Document(file_path)
        md_lines = []
        image_counter = 1

        for element in doc.element.body:
            if isinstance(element, CT_P):  # 段落处理
                paragraph_text = self._get_paragraph_text(element)
                if paragraph_text:
                    md_lines.append(self._format_paragraph(element, paragraph_text))
                
                # 图片处理
                image_counter = self._process_images(element, doc, image_dir, image_counter, md_lines)
                
            elif isinstance(element, CT_Tbl):  # 表格处理
                md_lines.extend(self._process_table(element, doc))

        return self._save_markdown(doc_dir, os.path.basename(file_path), md_lines)

    def _get_paragraph_text(self, element) -> str:
        """获取段落文本"""
        return element.text.strip()

    def _format_paragraph(self, element, text: str) -> str:
        """格式化段落为Markdown"""
        style = element.get("style", "")
        if "Heading" in style:
            level = int(style.replace("Heading ", ""))
            return f"{'#' * level} {text}\n\n"
        return f"{text}\n\n"

    def _process_images(self, element, doc, image_dir: str, counter: int, md_lines: List[str]) -> int:
        """处理段落中的图片"""
        for run in element.r_lst:
            blip = run.find(".//a:blip", namespaces={
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main"
            })
            if blip is not None:
                image_path = self._save_image(doc, blip, image_dir, counter)
                relative_path = os.path.join("images", os.path.basename(image_path))
                md_lines.append(f"![Image]({relative_path})\n\n")
                counter += 1
        return counter

    def _save_image(self, doc, blip, image_dir: str, counter: int) -> str:
        """保存图片到本地"""
        image_rel = blip.get(qn("r:embed"))
        image_part = doc.part.rels[image_rel].target_part
        image_ext = self._get_image_extension(image_part.content_type)
        image_path = os.path.join(image_dir, f"image_{counter}{image_ext}")
        
        with open(image_path, "wb") as f:
            f.write(image_part.blob)
            
        return image_path

    def _get_image_extension(self, content_type: str) -> str:
        """获取图片扩展名"""
        return {
            "image/png": ".png",
            "image/jpeg": ".jpg",
            "image/gif": ".gif",
            "image/bmp": ".bmp",
            "image/svg+xml": ".svg"
        }.get(content_type, ".png")

    def _process_table(self, element, doc) -> List[str]:
        """处理表格为Markdown格式"""
        table_lines = []
        # 通过docx库的Table对象处理
        for table in doc.tables:
            table_lines.append("\n")  # 表格前空行
            
            # 表头
            headers = [cell.text.strip() for cell in table.rows[0].cells]
            table_lines.append("| " + " | ".join(headers) + " |\n")
            table_lines.append("|" + " | ".join(["---"] * len(headers)) + "|\n")
            
            # 表格内容
            for row in table.rows[1:]:  # 跳过表头行
                row_data = [cell.text.strip() for cell in row.cells]
                table_lines.append("| " + " | ".join(row_data) + " |\n")
            
            table_lines.append("\n")  # 表格后空行
            
        return table_lines

    def _save_markdown(self, doc_dir: str, original_filename: str, content: List[str]) -> str:
        """保存Markdown文件"""
        md_filename = f"{os.path.splitext(original_filename)[0]}.md"
        md_path = os.path.join(doc_dir, md_filename)
        
        with open(md_path, "w", encoding="utf-8") as f:
            f.writelines(content)
            
        print(f"✅ Markdown文件已生成: {md_path}")
        return md_path


class ExcelParser(DocumentParser):
    def parse(self, file_path: str, output_dir: str) -> Optional[str]:
        """
        解析 Excel（支持 .xls 和 .xlsx），转换为 Markdown，并存放到结构化目录
        
        Args:
            file_path: Excel 文件路径
            output_dir: 输出目录路径
            
        Returns:
            str: 生成的 Markdown 文件路径，失败时返回 None
        """
        try:
            # 验证输入文件
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"输入文件不存在: {file_path}")
                
            # 创建输出目录
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            doc_dir = os.path.join(output_dir, file_name)
            os.makedirs(doc_dir, exist_ok=True)

            # 统一处理为.xlsx格式
            xlsx_path = self._convert_to_xlsx_if_needed(file_path, doc_dir)

            # 处理xlsx文件
            self._parse_xlsx(xlsx_path, doc_dir)

            # 添加返回语句
            md_path = os.path.join(doc_dir, file_name + ".md")
            return md_path
            
                
        except Exception as e:
            print(f"❌ 解析失败: {str(e)}")
            return None

    def _convert_to_xlsx_if_needed(self, file_path: str, output_dir: str) -> str:
        """将.xls转换为.xlsx（如需要）"""
        output_path = os.path.join(output_dir, os.path.basename(file_path).split(".")[0]) + '.xlsx'
        
        if file_path.lower().endswith('.xls'):
            try:
                user_install = f"file:///tmp/LibreOfficeProfile_{uuid.uuid4()}"
                subprocess.run([
                    "soffice", "--headless", f"-env:UserInstallation={user_install}", "--convert-to", "xlsx",
                    "--outdir", output_dir, file_path
                ], check=True, timeout=30)
                print(f"✅ 已转换.xls到.xlsx: {output_path}")
            except subprocess.TimeoutExpired:
                raise RuntimeError("文档转换超时，请确保已安装LibreOffice")
            except Exception as e:
                raise RuntimeError(f"文档转换失败: {str(e)}")
        else:
            shutil.copy2(file_path, output_path)
            
        return output_path
    def _parse_xlsx(self, file_path: str, doc_dir: str) -> str:
        """解析 .xlsx 并转换为 Markdown"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            all_md_lines = []
            
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                md_lines = [f"\n## {sheet_name}\n\n"]
                
                # 处理表头
                headers = self._get_headers(ws)
                md_lines.append(self._create_markdown_table_row(headers))
                md_lines.append(self._create_markdown_separator(len(headers)))
                
                # 处理数据行（包括合并单元格）
                merged_ranges = ws.merged_cells.ranges
                for row in ws.iter_rows(min_row=2):
                    row_data = self._process_row(row, merged_ranges)
                    md_lines.append(self._create_markdown_table_row(row_data))
                
                all_md_lines.extend(md_lines)
            
            # 保存 Markdown 文件
            md_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}.md"
            md_path = os.path.join(doc_dir, md_filename)
            
            with open(md_path, 'w', encoding='utf-8') as f:
                f.writelines(all_md_lines)
                
            print(f"✅ 解析完成: {md_path}")
            return md_path
            
        except Exception as e:
            raise RuntimeError(f"解析XLSX文件失败: {str(e)}")

    def _parse_xls(self, file_path: str, doc_dir: str) -> str:
        """解析 .xls 文件（待实现）"""
        raise NotImplementedError(".xls 解析功能暂未实现")

    # 辅助方法 ------------------------------------------------------
    
    def _get_headers(self, worksheet) -> List[str]:
        """获取表头行"""
        return [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
    
    def _process_row(self, row, merged_ranges) -> List[str]:
        """处理数据行，处理合并单元格"""
        row_data = []
        for cell in row:
            cell_value = self._get_cell_value(cell, merged_ranges)
            row_data.append(str(cell_value).replace('\n', '<br>'))  # 处理换行符
        return row_data
    
    def _get_cell_value(self, cell, merged_ranges):
        """获取单元格值，处理合并单元格"""
        for merged_range in merged_ranges:
            if (merged_range.min_row <= cell.row <= merged_range.max_row and
                merged_range.min_col <= cell.column <= merged_range.max_col):
                # 如果是合并区域的主单元格
                if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                    return cell.value
                # 如果是合并区域的其他单元格，返回主单元格的值
                return merged_range.start_cell.value
        return cell.value if cell.value is not None else ""
    
    def _create_markdown_table_row(self, cells: List[str]) -> str:
        """创建Markdown表格行"""
        return "| " + " | ".join(cells) + " |\n"
    
    def _create_markdown_separator(self, num_columns: int) -> str:
        """创建Markdown表格分隔线"""
        return "| " + " | ".join(["---"] * num_columns) + " |\n"

class HTMLParser(DocumentParser):
    """HTML文档解析器"""
    
    def parse(self, file_path: str, output_dir: str) -> Optional[str]:
        """
        解析 HTML 文件，转换为 Markdown
        
        Args:
            file_path: HTML 文件路径
            output_dir: 输出目录路径
            
        Returns:
            str: 生成的 Markdown 文件路径，失败时返回 None
        """
        try:
            # 验证输入文件
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"输入文件不存在: {file_path}")
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.html', '.htm']:
                raise ValueError(f"不支持的文件格式: {file_ext}")
            
            # 创建输出目录
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            doc_dir = os.path.join(output_dir, file_name)
            image_dir = os.path.join(doc_dir, "images")
            os.makedirs(image_dir, exist_ok=True)
            
            # 读取HTML内容
            html_content = read_text_file_with_encoding_detection(file_path)
            
            # 转换为Markdown
            md_content = self._html_to_markdown(html_content, file_path, image_dir)
            
            # 保存Markdown文件
            md_filename = f"{file_name}.md"
            md_path = os.path.join(doc_dir, md_filename)
            
            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            
            logger.info(f"✅ HTML解析完成: {md_path}")
            return md_path
            
        except Exception as e:
            logger.error(f"❌ HTML解析失败: {str(e)}")
            return None
    
    def _html_to_markdown(self, html_content: str, source_path: str, image_dir: str) -> str:
        """
        将HTML内容转换为Markdown格式
        
        Args:
            html_content: HTML内容字符串
            source_path: 源HTML文件路径（用于解析相对路径的图片）
            image_dir: 图片保存目录
            
        Returns:
            str: Markdown格式的内容
        """
        try:
            from bs4 import BeautifulSoup
            import html2text
            
            # 使用BeautifulSoup解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 处理图片 - 下载或复制到本地
            self._process_html_images(soup, source_path, image_dir)
            
            # 移除script和style标签
            for tag in soup(['script', 'style', 'meta', 'link']):
                tag.decompose()
            
            # 使用html2text转换为Markdown
            h = html2text.HTML2Text()
            h.ignore_links = False  # 保留链接
            h.ignore_images = False  # 保留图片
            h.ignore_emphasis = False  # 保留强调
            h.body_width = 0  # 不限制行宽
            h.unicode_snob = True  # 使用Unicode字符
            h.skip_internal_links = False  # 不跳过内部链接
            h.inline_links = True  # 使用内联链接格式
            h.protect_links = True  # 保护链接
            h.mark_code = True  # 标记代码块
            
            # 转换HTML为Markdown
            markdown_content = h.handle(str(soup))
            
            # 清理多余的空行
            markdown_content = re.sub(r'\n\s*\n\s*\n', '\n\n', markdown_content)
            
            return markdown_content.strip()
            
        except ImportError as e:
            logger.error("缺少必要的库，请安装: pip install beautifulsoup4 html2text")
            raise ImportError("请安装 beautifulsoup4 和 html2text 库") from e
        except Exception as e:
            logger.error(f"HTML转Markdown失败: {e}")
            raise
    
    def _process_html_images(self, soup, source_path: str, image_dir: str) -> None:
        """
        处理HTML中的图片，下载或复制到本地
        
        Args:
            soup: BeautifulSoup对象
            source_path: 源HTML文件路径
            image_dir: 图片保存目录
        """
        image_counter = 1
        source_dir = os.path.dirname(os.path.abspath(source_path))
        
        for img_tag in soup.find_all('img'):
            try:
                img_src = img_tag.get('src', '')
                if not img_src:
                    continue
                
                # 处理不同类型的图片源
                local_img_path = None
                
                # 处理base64图片
                if img_src.startswith('data:image'):
                    local_img_path = self._save_base64_image(img_src, image_dir, image_counter)
                
                # 处理HTTP/HTTPS图片
                elif img_src.startswith(('http://', 'https://')):
                    local_img_path = self._download_image(img_src, image_dir, image_counter)
                
                # 处理本地相对路径图片
                else:
                    abs_img_path = os.path.join(source_dir, img_src)
                    if os.path.exists(abs_img_path):
                        local_img_path = self._copy_local_image(abs_img_path, image_dir, image_counter)
                
                # 更新图片标签
                if local_img_path:
                    relative_path = os.path.join("images", os.path.basename(local_img_path))
                    img_tag['src'] = relative_path
                    image_counter += 1
                    
            except Exception as e:
                logger.warning(f"处理图片失败 {img_src}: {e}")
                continue
    
    def _save_base64_image(self, data_url: str, image_dir: str, counter: int) -> Optional[str]:
        """保存base64编码的图片"""
        try:
            import base64
            
            # 解析data URL
            header, encoded = data_url.split(',', 1)
            
            # 获取图片格式
            img_format = 'png'
            if 'jpeg' in header or 'jpg' in header:
                img_format = 'jpg'
            elif 'gif' in header:
                img_format = 'gif'
            elif 'webp' in header:
                img_format = 'webp'
            
            # 解码并保存
            img_data = base64.b64decode(encoded)
            img_path = os.path.join(image_dir, f"image_{counter}.{img_format}")
            
            with open(img_path, 'wb') as f:
                f.write(img_data)
            
            logger.debug(f"保存base64图片: {img_path}")
            return img_path
            
        except Exception as e:
            logger.warning(f"保存base64图片失败: {e}")
            return None
    
    def _download_image(self, url: str, image_dir: str, counter: int) -> Optional[str]:
        """下载网络图片"""
        try:
            response = requests.get(url, timeout=10, stream=True)
            response.raise_for_status()
            
            # 推断图片格式
            content_type = response.headers.get('content-type', '')
            img_format = 'png'
            if 'jpeg' in content_type or 'jpg' in content_type:
                img_format = 'jpg'
            elif 'gif' in content_type:
                img_format = 'gif'
            elif 'webp' in content_type:
                img_format = 'webp'
            
            img_path = os.path.join(image_dir, f"image_{counter}.{img_format}")
            
            with open(img_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            logger.debug(f"下载图片成功: {img_path}")
            return img_path
            
        except Exception as e:
            logger.warning(f"下载图片失败 {url}: {e}")
            return None
    
    def _copy_local_image(self, src_path: str, image_dir: str, counter: int) -> Optional[str]:
        """复制本地图片"""
        try:
            img_ext = os.path.splitext(src_path)[1] or '.png'
            img_path = os.path.join(image_dir, f"image_{counter}{img_ext}")
            
            shutil.copy2(src_path, img_path)
            logger.debug(f"复制本地图片: {img_path}")
            return img_path
            
        except Exception as e:
            logger.warning(f"复制本地图片失败 {src_path}: {e}")
            return None

class PDFConverter:
    """PDF转换器"""
    
    OFFICE_CONV_TIMEOUT = 60

    @staticmethod
    def _convert_using_libreoffice(input_path: str, output_path: str) -> None:
        """使用LibreOffice进行文件转换"""
        try:
            input_path = Path(input_path).resolve()
            output_path = Path(output_path).resolve()
            output_dir = output_path.parent
            output_dir.mkdir(parents=True, exist_ok=True)

            cmd = [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(input_path)
            ]
            
            logger.info(f"执行LibreOffice转换: {' '.join(cmd)}")
            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=PDFConverter.OFFICE_CONV_TIMEOUT
            )

            # 处理生成的文件路径
            input_basename = input_path.stem
            generated_pdf = output_dir / f"{input_basename}.pdf"

            # 重命名文件到指定路径
            if generated_pdf != output_path:
                shutil.move(str(generated_pdf), str(output_path))
                
            logger.info(f"LibreOffice转换成功: {output_path}")

        except subprocess.TimeoutExpired:
            raise RuntimeError(f"LibreOffice转换超时（{PDFConverter.OFFICE_CONV_TIMEOUT}秒）")
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr if e.stderr else "未知错误"
            raise RuntimeError(f"LibreOffice转换失败: {error_msg}")
        except FileNotFoundError:
            raise SystemExit("请先安装LibreOffice")

    @staticmethod
    def doc_to_pdf(input_path: str, output_path: str) -> None:
        """Word转PDF"""
        if sys.platform == "win32":
            try:
                from docx2pdf import convert
                convert(input_path, output_path)
                logger.info(f"Word转换成功: {output_path}")
            except ImportError:
                raise SystemExit("请安装 docx2pdf 库：pip install docx2pdf")
            except Exception as e:
                raise RuntimeError(f"Word转换失败: {e}")
        else:
            PDFConverter._convert_using_libreoffice(input_path, output_path)

    @staticmethod
    def excel_to_pdf(input_path: str, output_path: str) -> None:
        """Excel转PDF"""
        if sys.platform == "win32":
            try:
                from win32com import client
                excel = client.Dispatch("Excel.Application")
                excel.Visible = False
                wb = excel.Workbooks.Open(os.path.abspath(input_path))
                wb.ExportAsFixedFormat(0, os.path.abspath(output_path))
                wb.Close()
                excel.Quit()
                logger.info(f"Excel转换成功: {output_path}")
            except ImportError:
                raise SystemExit("请安装 pywin32 库：pip install pywin32")
            except Exception as e:
                raise RuntimeError(f"Excel转换失败: {e}")
        else:
            PDFConverter._convert_using_libreoffice(input_path, output_path)

    @staticmethod
    def image_to_pdf(input_path: str, output_path: str) -> None:
        """图片转PDF"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            image = Image.open(input_path)
            if image.mode != 'RGB':
                image = image.convert('RGB')

            img_width, img_height = image.size
            pdf_width, pdf_height = letter
            aspect = img_height / img_width

            # 智能缩放
            if img_width > pdf_width or img_height > pdf_height:
                if aspect >= 1:  # 竖版
                    new_height = pdf_height
                    new_width = new_height / aspect
                else:  # 横版
                    new_width = pdf_width
                    new_height = new_width * aspect
            else:
                new_width, new_height = img_width, img_height

            # 创建PDF
            c = canvas.Canvas(output_path, pagesize=letter)
            c.drawImage(
                input_path,
                (pdf_width - new_width) / 2,
                (pdf_height - new_height) / 2,
                width=new_width,
                height=new_height,
                preserveAspectRatio=True,
                mask='auto'
            )
            c.save()
            logger.info(f"图片转换成功: {output_path}")

        except ImportError:
            raise SystemExit("请安装 reportlab 库：pip install reportlab")   
        except Exception as e:
            raise RuntimeError(f"图片转换失败: {e}")

    @staticmethod
    def _ppt_to_pdf_windows(input_path: str, output_path: str) -> None:
        """使用PowerPoint COM组件转换（Windows）"""
        try:
            import pythoncom
            from win32com import client
            
            pythoncom.CoInitialize()
            
            try:
                ppt_app = client.Dispatch("PowerPoint.Application")
                ppt_app.Visible = False
                
                presentation = ppt_app.Presentations.Open(str(input_path), ReadOnly=True)
                
                try:
                    presentation.ExportAsFixedFormat(str(output_path), 2)  # ppFixedFormatTypePDF = 2
                    logger.info(f"PowerPoint转换成功: {output_path}")
                finally:
                    presentation.Close()
                    
            finally:
                try:
                    ppt_app.Quit()
                except:
                    pass
                pythoncom.CoUninitialize()
                
        except ImportError:
            raise SystemExit("Windows平台需要安装 pywin32 库：pip install pywin32")
        except Exception as e:
            raise RuntimeError(f"PowerPoint转换失败: {e}")

    @staticmethod
    def _ppt_to_pdf_libreoffice(input_path: str, output_path: str) -> None:
        """使用LibreOffice转换PPT"""
        try:
            input_path = Path(input_path).resolve()
            output_path = Path(output_path).resolve()
            
            temp_dir = output_path.parent / f"temp_ppt_conv_{uuid.uuid4()}"
            temp_dir.mkdir(exist_ok=True)
            temp_manager.add_temp_dir(str(temp_dir))
            
            try:
                if sys.platform == "win32":
                    soffice_cmd = "soffice"
                    user_install = f"file:///{str(Path.home() / 'temp' / f'LibreOfficeProfile_{uuid.uuid4()}').replace(chr(92), '/')}"
                else:
                    soffice_cmd = "soffice"
                    user_install = f"file:///tmp/LibreOfficeProfile_{uuid.uuid4()}"
                
                subprocess.run([
                    soffice_cmd, "--headless", 
                    f"-env:UserInstallation={user_install}", 
                    "--convert-to", "pdf",
                    "--outdir", str(temp_dir), 
                    str(input_path)
                ], check=True, timeout=PDFConverter.OFFICE_CONV_TIMEOUT,
                   stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                
                pdf_files = list(temp_dir.glob("*.pdf"))
                if not pdf_files:
                    raise FileNotFoundError("PDF转换结果文件未生成")
                
                generated_pdf = pdf_files[0]
                shutil.move(str(generated_pdf), str(output_path))
                logger.info(f"PPT转换成功: {output_path}")
                
            finally:
                if temp_dir.exists():
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"PPT转换超时（{PDFConverter.OFFICE_CONV_TIMEOUT}秒）")
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr.decode(errors='ignore')[:500] if e.stderr else "未知错误"
            raise RuntimeError(f"PPT转换失败: {error_msg}")
        except FileNotFoundError as e:
            if "LibreOffice" in str(e):
                raise SystemExit("请先安装LibreOffice")
            raise

    @staticmethod
    def ppt_to_pdf(input_path: str, output_path: str) -> None:
        """PPT转PDF"""
        try:
            input_path = Path(input_path).resolve()
            output_path = Path(output_path).resolve()
            
            if not input_path.exists():
                raise FileNotFoundError(f"输入文件不存在: {input_path}")
            
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            if sys.platform == "win32":
                try:
                    PDFConverter._ppt_to_pdf_windows(str(input_path), str(output_path))
                except (SystemExit, ImportError, RuntimeError) as e:
                    if "pywin32" not in str(e):
                        logger.warning(f"PowerPoint转换失败，尝试使用LibreOffice: {e}")
                    PDFConverter._ppt_to_pdf_libreoffice(str(input_path), str(output_path))
            else:
                PDFConverter._ppt_to_pdf_libreoffice(str(input_path), str(output_path))
                
        except Exception as e:
            raise RuntimeError(f"PPT转换失败: {e}")

    @staticmethod
    def ebook_to_pdf(input_path: str, output_path: str) -> None:
        """电子书转PDF"""
        try:
            input_path = Path(input_path).resolve()
            output_path = Path(output_path).resolve()
            
            if not input_path.exists():
                raise FileNotFoundError(f"文件不存在: {input_path}")
            
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            cmd = [
                'ebook-convert',
                str(input_path),
                str(output_path),
                '--pdf-page-numbers',
                '--preserve-cover-aspect-ratio',
                '--pdf-page-margin-bottom', '36',
                '--pdf-page-margin-top', '36',
                '--pdf-page-margin-left', '36',
                '--pdf-page-margin-right', '36',
                '--base-font-size', '12'
            ]
            
            result = subprocess.run(
                cmd, 
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=180
            )
            logger.info(f"电子书转换成功: {output_path}")
            
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr if e.stderr else "转换失败"
            raise RuntimeError(f"电子书转换失败: {error_msg}")
        except subprocess.TimeoutExpired:
            raise RuntimeError("电子书转换超时")
        except FileNotFoundError:
            raise SystemExit("请先安装 Calibre, sudo apt update, sudo apt install calibre")

    @staticmethod
    def convert_to_pdf(input_path: str, output_path: Optional[str] = None) -> str:
        """自动转换文件为PDF"""
        try:
            input_path = Path(input_path).resolve()
            if not input_path.exists():
                raise FileNotFoundError(f"文件不存在: {input_path}")
                
            if not output_path:
                output_path = input_path.with_suffix('.pdf')
            else:
                output_path = Path(output_path).resolve()

            ext = input_path.suffix.lower()
            output_path.parent.mkdir(parents=True, exist_ok=True)

            conversion_map = {
                ('.doc', '.docx'): PDFConverter.doc_to_pdf,
                ('.xls', '.xlsx'): PDFConverter.excel_to_pdf,
                ('.jpg', '.jpeg', '.png', '.bmp', '.svg', '.tif', '.gif', '.pcx', 
                 '.tga', '.exif', '.fpx', '.psd', '.cdr', '.pcd', '.dxf', '.ufo', 
                 '.eps', '.ai', '.raw', '.wmf', '.webp', '.avif', '.agnp'): PDFConverter.image_to_pdf,
                ('.ppt', '.pptx'): PDFConverter.ppt_to_pdf,
                ('.mobi', '.epub', '.azw', '.azw3', '.fb2', '.lit', '.lrf', '.pdb'): PDFConverter.ebook_to_pdf,
            }

            for extensions, converter_func in conversion_map.items():
                if ext in extensions:
                    with timer(f"{ext} 转 PDF"):
                        converter_func(str(input_path), str(output_path))
                    temp_manager.add_temp_file(str(output_path))
                    return str(output_path)
            
            raise ValueError(f"不支持的文件格式: {ext}")
            
        except Exception as e:
            logger.error(f"转换失败: {e}")
            raise


def process_excel_file(file_path: str, output_dir: str) -> str:
    """处理Excel文件"""
    try:
        output_md_path = os.path.join(output_dir, f"{Path(file_path).stem}.md")
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        excel_file = pd.ExcelFile(file_path)
        
        with open(output_md_path, "w", encoding="utf-8") as f:
            for i, sheet_name in enumerate(excel_file.sheet_names):
                try:
                    df = excel_file.parse(sheet_name)
                    markdown_table = df.to_markdown(index=False)
                    
                    f.write(f"# 工作表名称: {sheet_name}\n\n")
                    f.write(markdown_table)
                    
                    if i < len(excel_file.sheet_names) - 1:
                        f.write("\n\n---\n\n")
                        
                except Exception as e:
                    logger.warning(f"处理工作表 {sheet_name} 失败: {e}")
                    f.write(f"# 工作表名称: {sheet_name}\n\n")
                    f.write(f"处理此工作表时出错: {e}\n\n")
        
        logger.info(f"Excel文件处理完成: {output_md_path}")
        return output_md_path
        
    except Exception as e:
        logger.error(f"Excel文件处理失败: {e}")
        raise


def process_text_file(file_path: str, output_dir: str) -> str:
    """处理文本文件"""
    try:
        output_md_path = os.path.join(output_dir, f"{Path(file_path).stem}.md")
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 使用编码检测读取文件
        txt_content = read_text_file_with_encoding_detection(file_path)
        
        with open(output_md_path, "w", encoding="utf-8") as f:
            # 添加文件头信息
            f.write(f"# 文本文件: {Path(file_path).name}\n\n")
            f.write("```\n")
            f.write(txt_content)
            f.write("\n```\n")
        
        logger.info(f"文本文件处理完成: {output_md_path}")
        return output_md_path
        
    except Exception as e:
        logger.error(f"文本文件处理失败: {e}")
        raise


def clean_markdown_content(file_path: str) -> None:
    """清理Markdown文件中的图片标签"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 使用正则表达式去除图片标签
        cleaned_content = re.sub(r'!\[.*?\]\(.*?\)', '', content)
        
        # 清理多余的空行
        cleaned_content = re.sub(r'\n\s*\n\s*\n', '\n\n', cleaned_content)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned_content)
            
        logger.info(f"Markdown内容清理完成: {file_path}")
        
    except Exception as e:
        logger.error(f"清理Markdown内容失败: {e}")
        raise
    

def main():
    """主函数"""
    try:
        # 解析命令行参数
        parser = argparse.ArgumentParser(description="文件下载和解析工具")
        # parser.add_argument("--fileUrl", required=True, help="文件的 URL 地址")
        parser.add_argument("--config", default="config.ini", help="配置文件路径")
        parser.add_argument("--lang", default="ch", help="解析语言")
        parser.add_argument("--backend", default="vlm-vllm-engine", help="解析后端")
        parser.add_argument("--method", default="vlm", help="解析方法")
        parser.add_argument("--start-page", type=int, default=0, help="起始页码")
        parser.add_argument("--end-page", type=int, help="结束页码")
        args = parser.parse_args()

        # 初始化配置
        config = Config(args.config)
        
        # 设置环境变量
        os.environ['MINERU_MODEL_SOURCE'] = "modelscope"

        # 初始化docx解析器
        docx_parser = WordParser()

        # 初始化excel解析器
        excel_parser = ExcelParser()
        
        start_time = time.time()
        logger.info("开始处理文件...")

        #下载文件
        # with timer("文件下载"):
        #   downloader = FileDownloader(
        #      download_dir=config.get_download_dir(),
        #      max_size_mb=config.get_max_file_size_mb(),
        #      timeout=config.get_timeout_seconds()
        #   )
        # downloaded_file_path = downloader.download_file(args.fileUrl)
        # validate_file_size(downloaded_file_path, config.get_max_file_size_mb())

        # 测试文件用于模型下载 ====================================
        # "python script/convert_to_markdown.py --fileUrl https://typora-picture-room.oss-cn-chengdu.aliyuncs.com/img/8c4417320a8d303f547f5a49b7961d1.png"
        # downloaded_file_path = "/home/hp-2/Agent_Platform/MinerU/demo/pdfs/(2.16)第1章 概述1.ppt"
        # =======================================================

        # 获取文件列表
        files_path = "./temp"
        output_dir = config.get_output_dir()
        
        # 获取所有文件
        all_files = [f for f in os.listdir(files_path) if os.path.isfile(os.path.join(files_path, f))]
        total_files = len(all_files)
        
        if total_files == 0:
            logger.warning(f"文件夹 {files_path} 中没有找到文件")
            return 0
        
        logger.info(f"找到 {total_files} 个文件待处理")
        
        # 统计处理结果
        success_count = 0
        fail_count = 0
        results = []
        
        # 批量处理所有文件
        for index, file in enumerate(all_files, 1):
            downloaded_file_path = os.path.join(files_path, file)
            file_start_time = time.time()
            
            logger.info(f"[{index}/{total_files}] 开始处理: {file}")
            
            try:
                # 获取文件扩展名
                file_extension = Path(downloaded_file_path).suffix.lower()
                result_path = None

                # 处理Excel文件
                if file_extension in [".xls", ".xlsx"]:
                    try:
                        with timer("Excel文件处理"):
                            # result_path = process_excel_file(downloaded_file_path, output_dir)
                            temp_path = excel_parser.parse(downloaded_file_path, output_dir)

                            # 保留原始的后缀：xxx.md -> xxx.xlsx.md
                            original_filename = Path(downloaded_file_path).name
                            result_path = os.path.join(output_dir, original_filename + ".md")
                            shutil.move(temp_path, result_path)

                            # 删除temp_path文件夹
                            file_path = os.path.join(output_dir, original_filename.split(".")[0])
                            if os.path.exists(file_path):
                                shutil.rmtree(file_path)
                            
                            # 清理Markdown内容
                            with timer("内容清理"):
                                clean_markdown_content(result_path)

                    except ImportError as e:
                        logger.error(f"导入失败: {e}")
                        fail_count += 1
                        continue

                # 处理文本文件
                elif file_extension == ".txt":
                    with timer("文本文件处理"):
                        result_path = process_text_file(downloaded_file_path, output_dir)

                elif file_extension in [".doc", ".docx"]:
                    try:
                        with timer("Word文件处理"):
                            temp_path = docx_parser.parse(downloaded_file_path, output_dir)

                            # 保留原始的后缀：xxx.md -> xxx.docx.md
                            original_filename = Path(downloaded_file_path).name
                            result_path = os.path.join(output_dir, original_filename + ".md")
                            shutil.move(temp_path, result_path)

                            # 删除temp_path文件夹
                            file_path = os.path.join(output_dir, original_filename.split(".")[0])
                            if os.path.exists(file_path):
                                shutil.rmtree(file_path)

                            # 清理Markdown内容
                            with timer("内容清理"):
                                clean_markdown_content(result_path)

                    except Exception as e:
                        logger.error(f"Word文件处理失败: {e}")
                        fail_count += 1
                        continue

                # 在 main() 函数的文件类型判断部分添加:
                elif file_extension in [".html", ".htm"]:
                    try:
                        with timer("HTML文件处理"):
                            html_parser = HTMLParser()
                            temp_path = html_parser.parse(downloaded_file_path, output_dir)
                            
                            # 保留原始的后缀：xxx.md -> xxx.html.md
                            original_filename = Path(downloaded_file_path).name
                            result_path = os.path.join(output_dir, original_filename + ".md")
                            shutil.move(temp_path, result_path)
                            
                            # 删除temp_path文件夹
                            file_path = os.path.join(output_dir, original_filename.split(".")[0])
                            if os.path.exists(file_path):
                                shutil.rmtree(file_path)
                            
                            # 清理Markdown内容
                            with timer("内容清理"):
                                clean_markdown_content(result_path)
                                
                    except Exception as e:
                        logger.error(f"HTML文件处理失败: {e}")
                        fail_count += 1
                        continue

                # 处理PDF或其他可转换的文件
                else:
                    # 转换为PDF（如果需要）
                    pdf_path = downloaded_file_path
                    if file_extension != ".pdf":
                        with timer("文件格式转换"):
                            pdf_path = PDFConverter.convert_to_pdf(downloaded_file_path)

                    # 解析PDF文件
                    with timer("PDF解析"):
                        parse_doc(
                            path_list=[Path(pdf_path)], 
                            output_dir=output_dir, 
                            lang=args.lang,
                            backend=args.backend,
                            method=args.method,
                            start_page_id=args.start_page,
                            end_page_id=args.end_page
                        )

                    # 构建结果文件路径
                    name_without_suffix = Path(pdf_path).stem
                    # 获取原始文件名（包含后缀）
                    original_filename = Path(downloaded_file_path).name
                    result_file = f"{name_without_suffix}.md"
                    # 最终输出的文件名保留原后缀
                    final_result_file = f"{original_filename}.md"
                    result_path = os.path.join(output_dir, name_without_suffix, args.method, result_file)

                    if not os.path.exists(result_path):
                        raise FileNotFoundError(f"解析结果文件不存在: {result_path}")

                    # 移动result_path到output_dir，使用原始文件名
                    shutil.move(result_path, os.path.join(output_dir, final_result_file))
                    result_path = os.path.join(output_dir, final_result_file)

                    # 删除file_path文件夹
                    file_path = os.path.join(output_dir, name_without_suffix)
                    if os.path.exists(file_path):
                        shutil.rmtree(file_path)

                    # 清理Markdown内容
                    with timer("内容清理"):
                        clean_markdown_content(result_path)

                file_time = time.time() - file_start_time
                logger.info(f"[{index}/{total_files}] 处理完成: {file} (耗时: {file_time:.2f}秒)")
                
                if result_path:
                    results.append(result_path)
                    success_count += 1
                    print(f"✓ {result_path}")
                    
            except Exception as e:
                fail_count += 1
                logger.error(f"[{index}/{total_files}] 处理失败: {file} - {e}")
                print(f"✗ {file}: {e}")
                continue

        # 输出汇总信息
        total_time = time.time() - start_time
        logger.info("=" * 50)
        logger.info("批量处理完成!")
        logger.info(f"总文件数: {total_files}")
        logger.info(f"成功: {success_count}, 失败: {fail_count}")
        logger.info(f"总处理时间: {total_time:.2f}秒")
        logger.info("=" * 50)
        
        # 打印所有成功的结果路径
        print(f"\n处理结果汇总: 成功 {success_count}/{total_files}")
        for r in results:
            print(r)
            
        return 0 if fail_count == 0 else 1

    except KeyboardInterrupt:
        logger.info("用户中断操作")
        return 130
    except FileNotFoundError as e:
        logger.error(f"文件不存在: {e}")
        return 2
    except ValueError as e:
        logger.error(f"参数错误: {e}")
        return 22
    except PermissionError as e:
        logger.error(f"权限不足: {e}")
        return 13
    except RuntimeError as e:
        logger.error(f"运行时错误: {e}")
        return 1
    except SystemExit as e:
        logger.error(f"系统退出: {e}")
        return e.code if hasattr(e, 'code') else 1
    except Exception as e:
        logger.error(f"未知错误: {e}")
        logger.exception("详细错误信息:")
        return 1
    finally:
        # 清理临时文件
        temp_manager.cleanup()


if __name__ == "__main__":
    sys.exit(main())
